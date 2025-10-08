"""
导入管理器
支持从Excel、JSON文件导入账号和文章数据
"""
import logging
import json
from typing import Dict, List, Tuple, Any
from pathlib import Path
from datetime import datetime

logger = logging.getLogger(__name__)

# 检查openpyxl是否可用
try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl未安装，无法导入Excel文件")


class ImportManager:
    """导入管理器"""

    @staticmethod
    def import_from_json(
        file_path: str,
        account_manager,
        article_manager,
        skip_duplicates: bool = True
    ) -> Tuple[int, int, List[str]]:
        """
        从JSON文件导入数据

        Args:
            file_path: JSON文件路径
            account_manager: 账号管理器
            article_manager: 文章管理器
            skip_duplicates: 是否跳过重复数据（默认True）

        Returns:
            Tuple[int, int, List[str]]: (导入账号数, 导入文章数, 错误信息列表)
        """
        errors = []
        imported_accounts = 0
        imported_articles = 0

        try:
            # 读取JSON文件
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)

            # 验证数据格式
            if not isinstance(data, dict):
                return 0, 0, ["JSON格式错误：根节点必须是对象"]

            accounts = data.get('accounts', [])
            articles = data.get('articles', [])

            # 导入账号
            account_id_mapping = {}  # 旧ID -> 新ID的映射
            for account in accounts:
                try:
                    # 检查账号是否已存在
                    if skip_duplicates and account_manager.account_exists(account['name']):
                        # 获取现有账号ID
                        existing = account_manager.get_all_accounts()
                        for acc in existing:
                            if acc['name'] == account['name']:
                                account_id_mapping[account['id']] = acc['id']
                                break
                        continue

                    # 添加账号
                    new_id = account_manager.add_account(
                        name=account['name'],
                        category=account.get('category', ''),
                        description=account.get('description', ''),
                        avatar_url=account.get('avatar_url', '')
                    )

                    if new_id:
                        account_id_mapping[account['id']] = new_id
                        imported_accounts += 1

                except Exception as e:
                    errors.append(f"导入账号 '{account.get('name', '未知')}' 失败: {str(e)}")

            # 导入文章
            for article in articles:
                try:
                    # 获取对应的新账号ID
                    old_account_id = article.get('account_id')
                    new_account_id = account_id_mapping.get(old_account_id)

                    if not new_account_id:
                        errors.append(f"文章 '{article.get('title', '未知')}' 的账号不存在，已跳过")
                        continue

                    # 添加文章
                    article_id = article_manager.add_article(
                        account_id=new_account_id,
                        title=article.get('title', ''),
                        url=article.get('url', ''),
                        publish_date=article.get('publish_date'),
                        cover_image=article.get('cover_image', ''),
                        summary=article.get('summary', ''),
                        tags=article.get('tags', ''),
                        author=article.get('author', '')
                    )

                    if article_id:
                        imported_articles += 1
                    elif not skip_duplicates:
                        errors.append(f"文章 '{article.get('title', '未知')}' 导入失败（可能URL重复）")

                except Exception as e:
                    errors.append(f"导入文章 '{article.get('title', '未知')}' 失败: {str(e)}")

            logger.info(f"JSON导入完成: 账号{imported_accounts}个, 文章{imported_articles}篇")
            return imported_accounts, imported_articles, errors

        except FileNotFoundError:
            return 0, 0, [f"文件不存在: {file_path}"]
        except json.JSONDecodeError as e:
            return 0, 0, [f"JSON格式错误: {str(e)}"]
        except Exception as e:
            return 0, 0, [f"导入失败: {str(e)}"]

    @staticmethod
    def import_from_excel(
        file_path: str,
        account_manager,
        article_manager,
        skip_duplicates: bool = True
    ) -> Tuple[int, int, List[str]]:
        """
        从Excel文件导入数据

        Args:
            file_path: Excel文件路径
            account_manager: 账号管理器
            article_manager: 文章管理器
            skip_duplicates: 是否跳过重复数据（默认True）

        Returns:
            Tuple[int, int, List[str]]: (导入账号数, 导入文章数, 错误信息列表)
        """
        if not OPENPYXL_AVAILABLE:
            return 0, 0, ["openpyxl未安装，无法导入Excel文件"]

        errors = []
        imported_accounts = 0
        imported_articles = 0

        try:
            # 加载Excel文件
            wb = load_workbook(file_path, read_only=True)

            # 导入账号
            account_name_to_id = {}
            if '账号列表' in wb.sheetnames:
                ws = wb['账号列表']

                # 跳过标题行
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[0]:  # 如果账号名称为空，跳过
                        continue

                    try:
                        name = str(row[0])
                        category = str(row[1]) if row[1] else ''
                        description = str(row[4]) if len(row) > 4 and row[4] else ''
                        avatar_url = str(row[5]) if len(row) > 5 and row[5] else ''

                        # 检查是否已存在
                        if skip_duplicates and account_manager.account_exists(name):
                            # 获取现有账号ID
                            existing = account_manager.get_all_accounts()
                            for acc in existing:
                                if acc['name'] == name:
                                    account_name_to_id[name] = acc['id']
                                    break
                            continue

                        # 添加账号
                        new_id = account_manager.add_account(
                            name=name,
                            category=category,
                            description=description,
                            avatar_url=avatar_url
                        )

                        if new_id:
                            account_name_to_id[name] = new_id
                            imported_accounts += 1

                    except Exception as e:
                        errors.append(f"导入账号 '{row[0]}' 失败: {str(e)}")

            # 导入文章
            if '文章列表' in wb.sheetnames:
                ws = wb['文章列表']

                # 跳过标题行
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[1]:  # 如果标题为空，跳过
                        continue

                    try:
                        account_name = str(row[0]) if row[0] else ''
                        title = str(row[1])
                        url = str(row[2]) if row[2] else ''
                        publish_date = str(row[3]) if row[3] else None
                        author = str(row[4]) if len(row) > 4 and row[4] else ''
                        tags = str(row[5]) if len(row) > 5 and row[5] else ''
                        summary = str(row[6]) if len(row) > 6 and row[6] else ''

                        # 获取账号ID
                        account_id = account_name_to_id.get(account_name)
                        if not account_id:
                            errors.append(f"文章 '{title}' 的账号 '{account_name}' 不存在，已跳过")
                            continue

                        # 添加文章
                        article_id = article_manager.add_article(
                            account_id=account_id,
                            title=title,
                            url=url,
                            publish_date=publish_date,
                            cover_image='',
                            summary=summary,
                            tags=tags,
                            author=author
                        )

                        if article_id:
                            imported_articles += 1
                        elif not skip_duplicates:
                            errors.append(f"文章 '{title}' 导入失败（可能URL重复）")

                    except Exception as e:
                        errors.append(f"导入文章 '{row[1] if len(row) > 1 else '未知'}' 失败: {str(e)}")

            wb.close()
            logger.info(f"Excel导入完成: 账号{imported_accounts}个, 文章{imported_articles}篇")
            return imported_accounts, imported_articles, errors

        except FileNotFoundError:
            return 0, 0, [f"文件不存在: {file_path}"]
        except Exception as e:
            return 0, 0, [f"导入失败: {str(e)}"]
