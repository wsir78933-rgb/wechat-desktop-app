"""
导出管理器
负责数据导出功能：Excel、JSON、Markdown
"""
import logging
import json
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("openpyxl未安装，Excel导出功能将不可用")


logger = logging.getLogger(__name__)


class DateTimeEncoder(json.JSONEncoder):
    """自定义JSON编码器（处理日期时间）"""

    def default(self, obj):
        if isinstance(obj, datetime):
            return obj.strftime('%Y-%m-%d %H:%M:%S')
        return super().default(obj)


class ExportManager:
    """导出管理器"""

    @staticmethod
    def export_to_excel(
        accounts: List[Dict[str, Any]],
        articles: List[Dict[str, Any]],
        file_path: str
    ) -> bool:
        """
        导出为Excel文件

        Args:
            accounts: 账号列表
            articles: 文章列表
            file_path: 保存路径

        Returns:
            bool: 是否成功
        """
        if not OPENPYXL_AVAILABLE:
            logger.error("openpyxl未安装，无法导出Excel")
            return False

        try:
            wb = Workbook()

            # ========== 账号表 ==========
            ws_accounts = wb.active
            ws_accounts.title = "账号列表"

            # 设置表头
            account_headers = [
                "账号ID", "账号名称", "分类", "描述",
                "头像链接", "文章数", "最新文章日期", "创建时间"
            ]
            ws_accounts.append(account_headers)

            # 设置表头样式
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)

            for col_num, cell in enumerate(ws_accounts[1], 1):
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 填充账号数据
            for account in accounts:
                ws_accounts.append([
                    account.get('id', ''),
                    account.get('name', ''),
                    account.get('category', ''),
                    account.get('description', ''),
                    account.get('avatar_url', ''),
                    account.get('article_count', 0),
                    account.get('latest_date', ''),
                    account.get('created_at', '')
                ])

            # 设置列宽
            column_widths = {
                'A': 10,  # 账号ID
                'B': 20,  # 账号名称
                'C': 12,  # 分类
                'D': 40,  # 描述
                'E': 40,  # 头像链接
                'F': 10,  # 文章数
                'G': 15,  # 最新文章日期
                'H': 20   # 创建时间
            }

            for col, width in column_widths.items():
                ws_accounts.column_dimensions[col].width = width

            # ========== 文章表 ==========
            ws_articles = wb.create_sheet("文章列表")

            # 设置表头
            article_headers = [
                "文章ID", "账号名称", "账号分类", "文章标题",
                "文章链接", "发布日期", "作者", "标签", "摘要", "创建时间"
            ]
            ws_articles.append(article_headers)

            # 设置表头样式
            for col_num, cell in enumerate(ws_articles[1], 1):
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 填充文章数据
            for article in articles:
                ws_articles.append([
                    article.get('id', ''),
                    article.get('account_name', ''),
                    article.get('account_category', ''),
                    article.get('title', ''),
                    article.get('url', ''),
                    article.get('publish_date', ''),
                    article.get('author', ''),
                    article.get('tags', ''),
                    article.get('summary', ''),
                    article.get('created_at', '')
                ])

            # 设置列宽
            article_column_widths = {
                'A': 10,  # 文章ID
                'B': 20,  # 账号名称
                'C': 12,  # 账号分类
                'D': 40,  # 文章标题
                'E': 50,  # 文章链接
                'F': 12,  # 发布日期
                'G': 15,  # 作者
                'H': 20,  # 标签
                'I': 40,  # 摘要
                'J': 20   # 创建时间
            }

            for col, width in article_column_widths.items():
                ws_articles.column_dimensions[col].width = width

            # ========== 统计表 ==========
            ws_stats = wb.create_sheet("统计信息")

            stats_data = [
                ["统计项", "数值"],
                ["账号总数", len(accounts)],
                ["文章总数", len(articles)],
                ["导出时间", datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
            ]

            for row in stats_data:
                ws_stats.append(row)

            # 设置统计表样式
            for cell in ws_stats[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            ws_stats.column_dimensions['A'].width = 20
            ws_stats.column_dimensions['B'].width = 30

            # 保存文件
            Path(file_path).parent.mkdir(parents=True, exist_ok=True)
            wb.save(file_path)

            logger.info(f"Excel导出成功: {file_path}")
            logger.info(f"导出账号{len(accounts)}个, 文章{len(articles)}篇")
            return True

        except Exception as e:
            logger.error(f"Excel导出失败: {e}")
            return False

    @staticmethod
    def export_to_json(
        accounts: List[Dict[str, Any]],
        articles: List[Dict[str, Any]],
        file_path: str,
        indent: int = 2
    ) -> bool:
        """
        导出为JSON文件

        Args:
            accounts: 账号列表
            articles: 文章列表
            file_path: 保存路径
            indent: 缩进空格数（默认2）

        Returns:
            bool: 是否成功
        """
        try:
            data = {
                'metadata': {
                    'export_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'total_accounts': len(accounts),
                    'total_articles': len(articles),
                    'version': '1.0'
                },
                'accounts': accounts,
                'articles': articles
            }

            # 确保目录存在
            Path(file_path).parent.mkdir(parents=True, exist_ok=True)

            # 写入JSON文件
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=indent, cls=DateTimeEncoder)

            logger.info(f"JSON导出成功: {file_path}")
            logger.info(f"导出账号{len(accounts)}个, 文章{len(articles)}篇")
            return True

        except Exception as e:
            logger.error(f"JSON导出失败: {e}")
            return False

    @staticmethod
    def export_to_markdown(
        accounts: List[Dict[str, Any]],
        articles: List[Dict[str, Any]],
        file_path: str,
        group_by_account: bool = True
    ) -> bool:
        """
        导出为Markdown文件

        Args:
            accounts: 账号列表
            articles: 文章列表
            file_path: 保存路径
            group_by_account: 是否按账号分组（默认True）

        Returns:
            bool: 是否成功
        """
        try:
            # 确保目录存在
            Path(file_path).parent.mkdir(parents=True, exist_ok=True)

            with open(file_path, 'w', encoding='utf-8') as f:
                # 写入标题
                f.write("# 对标账号管理 - 数据导出\n\n")
                f.write(f"**导出时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
                f.write(f"**统计信息**: 账号 {len(accounts)} 个 | 文章 {len(articles)} 篇\n\n")
                f.write("---\n\n")

                if group_by_account:
                    # 按账号分组导出
                    f.write("## 账号与文章列表\n\n")

                    # 创建账号ID到文章的映射
                    account_articles_map = {}
                    for article in articles:
                        account_id = article.get('account_id')
                        if account_id not in account_articles_map:
                            account_articles_map[account_id] = []
                        account_articles_map[account_id].append(article)

                    # 遍历账号
                    for idx, account in enumerate(accounts, 1):
                        account_id = account.get('id')
                        account_name = account.get('name', '未命名账号')
                        category = account.get('category', '未分类')
                        description = account.get('description', '')
                        article_count = account.get('article_count', 0)

                        # 写入账号信息
                        f.write(f"### {idx}. {account_name}\n\n")
                        f.write(f"- **分类**: {category}\n")
                        f.write(f"- **文章数**: {article_count}\n")

                        if description:
                            f.write(f"- **描述**: {description}\n")

                        f.write("\n")

                        # 写入该账号的文章列表
                        account_articles = account_articles_map.get(account_id, [])

                        if account_articles:
                            f.write("#### 文章列表\n\n")

                            for art_idx, article in enumerate(account_articles, 1):
                                title = article.get('title', '未命名文章')
                                url = article.get('url', '')
                                publish_date = article.get('publish_date', '未知日期')
                                tags = article.get('tags', '')
                                summary = article.get('summary', '')

                                f.write(f"{art_idx}. **{title}**\n")
                                f.write(f"   - 链接: {url}\n")
                                f.write(f"   - 发布日期: {publish_date}\n")

                                if tags:
                                    f.write(f"   - 标签: {tags}\n")

                                if summary:
                                    f.write(f"   - 摘要: {summary}\n")

                                f.write("\n")
                        else:
                            f.write("*暂无文章*\n\n")

                        f.write("---\n\n")

                else:
                    # 不分组，分别列出账号和文章
                    f.write("## 账号列表\n\n")

                    # 写入账号表格
                    f.write("| 序号 | 账号名称 | 分类 | 文章数 | 描述 |\n")
                    f.write("|------|---------|------|--------|------|\n")

                    for idx, account in enumerate(accounts, 1):
                        name = account.get('name', '未命名账号')
                        category = account.get('category', '未分类')
                        article_count = account.get('article_count', 0)
                        description = account.get('description', '')[:50]  # 限制长度

                        f.write(f"| {idx} | {name} | {category} | {article_count} | {description} |\n")

                    f.write("\n---\n\n")

                    # 写入文章列表
                    f.write("## 文章列表\n\n")

                    f.write("| 序号 | 账号名称 | 文章标题 | 发布日期 | 链接 |\n")
                    f.write("|------|---------|---------|---------|------|\n")

                    for idx, article in enumerate(articles, 1):
                        account_name = article.get('account_name', '未知账号')
                        title = article.get('title', '未命名文章')
                        publish_date = article.get('publish_date', '未知日期')
                        url = article.get('url', '')

                        f.write(f"| {idx} | {account_name} | {title} | {publish_date} | [链接]({url}) |\n")

                    f.write("\n")

            logger.info(f"Markdown导出成功: {file_path}")
            logger.info(f"导出账号{len(accounts)}个, 文章{len(articles)}篇")
            return True

        except Exception as e:
            logger.error(f"Markdown导出失败: {e}")
            return False

    @staticmethod
    def export_accounts_only(
        accounts: List[Dict[str, Any]],
        file_path: str,
        format: str = 'json'
    ) -> bool:
        """
        仅导出账号数据

        Args:
            accounts: 账号列表
            file_path: 保存路径
            format: 导出格式，可选 'json', 'excel', 'markdown'

        Returns:
            bool: 是否成功
        """
        if format == 'json':
            return ExportManager.export_to_json(accounts, [], file_path)
        elif format == 'excel':
            return ExportManager.export_to_excel(accounts, [], file_path)
        elif format == 'markdown':
            return ExportManager.export_to_markdown(accounts, [], file_path, group_by_account=False)
        else:
            logger.error(f"不支持的导出格式: {format}")
            return False

    @staticmethod
    def export_articles_by_account(
        account: Dict[str, Any],
        articles: List[Dict[str, Any]],
        file_path: str,
        format: str = 'markdown'
    ) -> bool:
        """
        导出单个账号及其文章

        Args:
            account: 账号信息
            articles: 文章列表
            file_path: 保存路径
            format: 导出格式

        Returns:
            bool: 是否成功
        """
        if format == 'json':
            return ExportManager.export_to_json([account], articles, file_path)
        elif format == 'excel':
            return ExportManager.export_to_excel([account], articles, file_path)
        elif format == 'markdown':
            return ExportManager.export_to_markdown([account], articles, file_path, group_by_account=True)
        else:
            logger.error(f"不支持的导出格式: {format}")
            return False
